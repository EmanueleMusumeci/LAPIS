"""
Utility functions for saving and managing planning results.
"""

from pathlib import Path
import csv
from datetime import datetime
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import CellIsRule
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def save_plan_to_file(data_folder, problem_id, plan, is_valid, attempt_history, formula):
    """
    Save the generated plan (or failure indicator) to a file with full history.
    
    Args:
        data_folder: Path to the data_* folder
        problem_id: ID of the problem (folder name like '100', '101', etc.)
        plan: The generated plan (list or string)
        is_valid: Whether the plan is valid or failed
        attempt_history: List of dictionaries containing attempt history
        formula: The LTL formula used for verification
    """
    # Create results folder if it doesn't exist
    # New structure: data_folder / problem_id / 'high'
    results_folder = data_folder / problem_id / 'high'
    results_folder.mkdir(parents=True, exist_ok=True)
    
    # Create file path with problem_id as filename
    output_file = results_folder / 'plan.txt'
    
    # Write full history and final result
    with open(output_file, 'w') as f:
        f.write("=" * 80 + "\n")
        f.write(f"PROBLEM {problem_id} - GENERATION HISTORY\n")
        f.write("=" * 80 + "\n\n")
        
        f.write("LTL FORMULA:\n")
        f.write(f"{formula}\n\n")
        
        # Write each attempt's history
        for i, attempt in enumerate(attempt_history, 1):
            f.write(f"\n{'='*80}\n")
            f.write(f"ATTEMPT {i}\n")
            f.write(f"{'='*80}\n\n")
            
            # Write reasoning if available
            if 'reasoning' in attempt and attempt['reasoning']:
                f.write("REASONING:\n")
                f.write(f"{attempt['reasoning']}\n\n")
            
            f.write("GENERATED PLAN:\n")
            if isinstance(attempt['plan'], list):
                for j, step in enumerate(attempt['plan'], 1):
                    f.write(f"{j}. {step}\n")
            else:
                f.write(f"{attempt['plan']}\n")
            
            f.write(f"\nVALID: {attempt['valid']}\n")
            
            if not attempt['valid']:
                # Write trace
                if 'trace' in attempt and attempt['trace']:
                    f.write(f"\nTRACE (State Sequence):\n")
                    for state_idx, state in enumerate(attempt['trace']):
                        f.write(f"  State {state_idx}:\n")
                        for fact in sorted(state):
                            f.write(f"    - {fact}\n")
                
                # Write violation details
                if 'violation_details' in attempt and attempt['violation_details']:
                    violation_report = attempt['violation_details'].get('violation_report', 'N/A')
                    if violation_report != 'N/A':
                        f.write(f"\nVIOLATION DETAILS:\n")
                        f.write(f"{violation_report}\n")
                
                # Write feedback
                if 'feedback' in attempt and attempt['feedback']:
                    f.write(f"\nFEEDBACK PROVIDED:\n")
                    f.write(f"{attempt['feedback']}\n")
        
        # Write final result
        f.write(f"\n{'='*80}\n")
        f.write("FINAL RESULT\n")
        f.write(f"{'='*80}\n\n")
        
        if is_valid and plan:
            f.write("STATUS: SUCCESS\n\n")
            f.write("FINAL PLAN:\n")
            if isinstance(plan, list):
                for i, step in enumerate(plan, 1):
                    f.write(f"{i}. {step}\n")
            else:
                f.write(f"{plan}\n")
        else:
            f.write("STATUS: FAILED\n\n")
            f.write("failed plan\n")
    
    print(f'\n[Plan history saved to: {output_file}]')


def append_result_to_summary(data_folder, problem_id, is_valid, attempts):
    """
    Append a single result to the summary file.
    
    Args:
        data_folder: Path to the data_* folder
        problem_id: ID of the problem (folder name like '100', '101', etc.)
        is_valid: Whether the plan is valid or failed
        attempts: Number of attempts made
    """
    # Create results folder if it doesn't exist
    results_folder = data_folder / 'results'
    results_folder.mkdir(exist_ok=True)
    
    summary_file = results_folder / 'summary.txt'
    
    # Create status line
    status = "✓" if is_valid else "✗"
    result_line = f"  {status} Problem {problem_id}: {'Success' if is_valid else 'Failed'} (Attempts: {attempts})\n"
    
    # Check if file is new/empty and needs header
    write_header = not summary_file.exists() or summary_file.stat().st_size == 0
    
    # Append to summary file
    with open(summary_file, 'a') as f:
        if write_header:
            f.write("Detailed Results:\n")
        f.write(result_line)
    
    print(f'[Result appended to: {summary_file}]')


def append_to_csv_log(data_folder, problem_id, domain_name, num_constraints, 
                      domain, problem, actions, goal, constraints,
                      formula, formula_reasoning, fluent_syntax,
                      plan, plan_reasoning, trace, trace_reasoning,
                      is_valid, num_loops, status, feedback=None):
    """
    Append planning result to CSV log file.
    
    Args:
        data_folder: Path to the data_* folder
        problem_id: ID of the problem
        domain_name: Name of the domain
        num_constraints: Number of constraints
        domain: Domain description
        problem: Problem description
        actions: Available actions
        goal: Goal description
        constraints: Constraints description
        formula: LTL formula
        formula_reasoning: Reasoning for formula generation
        fluent_syntax: Fluent syntax
        plan: Generated plan
        plan_reasoning: Reasoning for plan generation
        trace: Generated trace
        trace_reasoning: Reasoning for trace generation
        is_valid: Whether the plan is valid
        num_loops: Number of attempts/loops
        status: 'completed' or 'failed'
        feedback: Optional feedback for failed plans
    """
    # Create results folder if it doesn't exist
    results_folder = data_folder / 'results'
    results_folder.mkdir(exist_ok=True)
    
    csv_file = results_folder / 'planning_log.csv'
    
    # Check if file exists to determine if we need to write header
    file_exists = csv_file.exists()
    
    # No truncation - include ALL data
    def truncate(text, max_length=None):
        return str(text)
    
    # Prepare row data with complete text (no truncation)
    row_data = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'problem_id': problem_id,
        'domain_name': domain_name,
        'num_constraints': num_constraints,
        'domain': str(domain),
        'problem': str(problem),
        'actions': str(actions),
        'goal': str(goal),
        'constraints': str(constraints),
        'ltl_formula': str(formula),
        'formula_reasoning': str(formula_reasoning),
        'fluent_syntax': str(fluent_syntax),
        'plan': str(plan),
        'plan_reasoning': str(plan_reasoning),
        'trace': str(trace),
        'trace_reasoning': str(trace_reasoning),
        'is_valid': is_valid,
        'num_loops': num_loops,
        'status': status,
        'feedback': str(feedback) if feedback else ''
    }
    
    # Write to CSV
    with open(csv_file, 'a', newline='', encoding='utf-8') as f:
        fieldnames = list(row_data.keys())
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        
        # Write header if file is new
        if not file_exists:
            writer.writeheader()
        
        writer.writerow(row_data)
    
    print(f'[Result appended to CSV: {csv_file}]')


def append_to_excel_log(data_folder, problem_id, domain_name, num_constraints, 
                        domain, problem, actions, goal, constraints,
                        formula, formula_reasoning, fluent_syntax,
                        plan, plan_reasoning, trace, trace_reasoning,
                        is_valid, num_loops, status, feedback=None):
    """
    Append planning result to Excel log file with colored rows.
    Green for successful plans, red for failed plans.
    
    Args:
        Same as append_to_csv_log
    """
    if not EXCEL_AVAILABLE:
        return  # Silently skip if openpyxl not available
    
    # Create results folder if it doesn't exist
    results_folder = data_folder / 'results'
    results_folder.mkdir(exist_ok=True)
    
    excel_file = results_folder / 'planning_log.xlsx'
    
    # Helper function to truncate long strings
    # No truncation - include ALL data
    def truncate(text, max_length=None):
        return str(text)
    
    # Prepare row data with complete text (no truncation)
    row_data = [
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        str(problem_id),
        str(domain_name),
        str(num_constraints),
        str(domain),
        str(problem),
        str(actions),
        str(goal),
        str(constraints),
        str(formula),
        str(formula_reasoning),
        str(fluent_syntax),
        str(plan),
        str(plan_reasoning),
        str(trace),
        str(trace_reasoning),
        is_valid,
        num_loops,
        status,
        truncate(feedback, 300) if feedback else ''
    ]
    
    headers = [
        'timestamp', 'problem_id', 'domain_name', 'num_constraints',
        'domain', 'problem', 'actions', 'goal', 'constraints',
        'ltl_formula', 'formula_reasoning', 'fluent_syntax',
        'plan', 'plan_reasoning', 'trace', 'trace_reasoning',
        'is_valid', 'num_loops', 'status', 'feedback'
    ]
    
    # Load or create workbook
    if excel_file.exists():
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Planning Log"
        # Add headers
        ws.append(headers)
        # Format header row
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Set column widths (fixed sizes for better readability)
        column_widths = {
            'A': 18,  # timestamp
            'B': 12,  # problem_id
            'C': 15,  # domain_name
            'D': 15,  # num_constraints
            'E': 30,  # domain
            'F': 30,  # problem
            'G': 30,  # actions
            'H': 25,  # goal
            'I': 30,  # constraints
            'J': 35,  # ltl_formula
            'K': 35,  # formula_reasoning
            'L': 30,  # fluent_syntax
            'M': 35,  # plan
            'N': 35,  # plan_reasoning
            'O': 35,  # trace
            'P': 35,  # trace_reasoning
            'Q': 10,  # is_valid
            'R': 12,  # num_loops
            'S': 12,  # status
            'T': 35   # feedback
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add conditional formatting rules for the entire sheet
        # Green for successful rows (where is_valid = TRUE)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_rule = CellIsRule(operator='equal', formula=['TRUE'], fill=green_fill)
        
        # Red for failed rows (where is_valid = FALSE)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_rule = CellIsRule(operator='equal', formula=['FALSE'], fill=red_fill)
        
        # Apply to column Q (is_valid) - this will be used as reference
        # We'll apply the formatting to entire rows later
    
    # Append data row
    ws.append(row_data)
    
    # Get the last row and apply formatting
    last_row = ws.max_row
    
    # Apply color based on success/failure
    if is_valid:
        # Green for success
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        # Red for failure
        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Apply fill and text wrapping to all cells in the row
    alignment = Alignment(wrap_text=True, vertical='top')
    for cell in ws[last_row]:
        cell.fill = fill
        cell.alignment = alignment
    
    # Save workbook
    wb.save(excel_file)
    print(f'[Result appended to Excel: {excel_file}]')
